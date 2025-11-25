#!/usr/bin/env python3
"""
GlamorByBee Master Data Generator
Generates Excel file with dummy data for testing
"""

import json
from datetime import datetime, timedelta
import random

# Try to import openpyxl, if not available, provide instructions
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl")
    print("Generating JSON data files instead...")
    EXCEL_AVAILABLE = False
else:
    EXCEL_AVAILABLE = True


# Sample data
FIRST_NAMES = ["Sarah", "Amaka", "Jennifer", "Emily", "Jessica", "Ashley", "Michelle", 
               "Stephanie", "Nicole", "Brittany", "Amanda", "Melissa", "Rebecca", "Laura"]
LAST_NAMES = ["Johnson", "Williams", "Davis", "Brown", "Miller", "Wilson", "Moore", 
              "Taylor", "Anderson", "Thomas", "Jackson", "White", "Harris", "Martin"]

SERVICES = [
    {"name": "Bridal Glam", "price": 300.00, "duration": 120},
    {"name": "Natural Beat", "price": 150.00, "duration": 60},
    {"name": "Evening Elegance", "price": 200.00, "duration": 90},
    {"name": "Special FX", "price": 250.00, "duration": 90},
    {"name": "Photo Ready", "price": 175.00, "duration": 75},
    {"name": "Glam Squad Package", "price": 400.00, "duration": 150},
    {"name": "Makeup Lesson", "price": 125.00, "duration": 60}
]

STREETS = ["Main St", "Oak Ave", "Maple Dr", "Cedar Ln", "Pine Rd", "Elm St", 
           "Washington Blvd", "Jefferson Ave", "Madison Dr", "Monroe St"]
CITIES = ["Dallas", "Plano", "Frisco", "McKinney", "Allen", "Richardson", "Carrollton"]

STATUSES = ["confirmed", "pending", "completed", "cancelled"]
PAYMENT_METHODS = ["cash", "credit_card", "debit_card", "venmo", "zelle"]


def generate_customers(count=50):
    """Generate customer data"""
    customers = []
    for i in range(1, count + 1):
        first_name = random.choice(FIRST_NAMES)
        last_name = random.choice(LAST_NAMES)
        customers.append({
            "customer_id": f"CUST{i:04d}",
            "first_name": first_name,
            "last_name": last_name,
            "email": f"{first_name.lower()}.{last_name.lower()}{i}@email.com",
            "phone": f"(214) {random.randint(100, 999)}-{random.randint(1000, 9999)}",
            "date_registered": (datetime.now() - timedelta(days=random.randint(1, 365))).strftime("%Y-%m-%d"),
            "total_visits": random.randint(1, 15),
            "loyalty_points": random.randint(0, 500),
            "preferred_contact": random.choice(["email", "phone", "text"]),
            "notes": random.choice(["VIP client", "Referred by friend", "Social media", "Walk-in", "Regular client", ""])
        })
    return customers


def generate_addresses(customers):
    """Generate address data linked to customers"""
    addresses = []
    for idx, customer in enumerate(customers, 1):
        # 80% of customers have home address, 20% have both home and billing
        addresses.append({
            "address_id": f"ADDR{idx:04d}",
            "customer_id": customer["customer_id"],
            "address_type": "home",
            "street_address": f"{random.randint(100, 9999)} {random.choice(STREETS)}",
            "apt_unit": random.choice(["", f"Apt {random.randint(1, 999)}", f"Suite {random.randint(100, 500)}"]),
            "city": random.choice(CITIES),
            "state": "TX",
            "zip_code": f"75{random.randint(0, 9)}{random.randint(10, 99)}",
            "is_default": True
        })
        
        # 20% have separate billing address
        if random.random() < 0.2:
            addresses.append({
                "address_id": f"ADDR{len(addresses) + 1:04d}",
                "customer_id": customer["customer_id"],
                "address_type": "billing",
                "street_address": f"{random.randint(100, 9999)} {random.choice(STREETS)}",
                "apt_unit": "",
                "city": random.choice(CITIES),
                "state": "TX",
                "zip_code": f"75{random.randint(0, 9)}{random.randint(10, 99)}",
                "is_default": False
            })
    
    return addresses


def generate_appointments(customers, count=100):
    """Generate appointment data"""
    appointments = []
    base_date = datetime.now() - timedelta(days=90)
    
    for i in range(1, count + 1):
        customer = random.choice(customers)
        service = random.choice(SERVICES)
        appointment_date = base_date + timedelta(days=random.randint(0, 120))
        hour = random.randint(7, 18)
        
        # Determine status based on date
        if appointment_date < datetime.now() - timedelta(days=1):
            status = random.choice(["completed", "completed", "completed", "cancelled"])
        elif appointment_date < datetime.now():
            status = "completed"
        else:
            status = random.choice(["confirmed", "confirmed", "pending"])
        
        appointments.append({
            "appointment_id": f"APT{i:04d}",
            "customer_id": customer["customer_id"],
            "service_name": service["name"],
            "appointment_date": appointment_date.strftime("%Y-%m-%d"),
            "appointment_time": f"{hour:02d}:00",
            "duration_minutes": service["duration"],
            "location": random.choice(["studio", "studio", "studio", "home"]),
            "status": status,
            "deposit_paid": service["price"] * 0.30,
            "total_amount": service["price"],
            "balance_due": service["price"] * 0.70 if status != "completed" else 0,
            "payment_method": random.choice(PAYMENT_METHODS) if status == "completed" else "",
            "notes": random.choice(["", "Allergic to latex", "Brings own brushes", "Prefers natural look", "Rush service"])
        })
    
    return appointments


def generate_products():
    """Generate product/service catalog"""
    products = []
    for idx, service in enumerate(SERVICES, 1):
        products.append({
            "product_id": f"PROD{idx:03d}",
            "service_name": service["name"],
            "category": random.choice(["Bridal", "Special Event", "Everyday", "Educational"]),
            "base_price": service["price"],
            "duration_minutes": service["duration"],
            "deposit_required": service["price"] * 0.30,
            "is_active": True,
            "description": f"Professional {service['name'].lower()} service",
            "includes_lashes": random.choice([True, False]),
            "includes_touch_up": random.choice([True, False])
        })
    
    # Add product sales items
    makeup_products = [
        {"name": "Lash Set - Natural", "price": 25.00, "category": "Lashes"},
        {"name": "Lash Set - Dramatic", "price": 35.00, "category": "Lashes"},
        {"name": "Lash Glue - Sensitive", "price": 15.00, "category": "Lashes"},
        {"name": "Setting Spray", "price": 30.00, "category": "Products"},
        {"name": "Makeup Remover Wipes", "price": 12.00, "category": "Products"},
        {"name": "Touch-Up Kit", "price": 45.00, "category": "Products"},
    ]
    
    for idx, product in enumerate(makeup_products, len(products) + 1):
        products.append({
            "product_id": f"PROD{idx:03d}",
            "service_name": product["name"],
            "category": product["category"],
            "base_price": product["price"],
            "duration_minutes": 0,
            "deposit_required": 0,
            "is_active": True,
            "description": f"{product['name']} for retail",
            "includes_lashes": False,
            "includes_touch_up": False
        })
    
    return products


def generate_invoices(appointments):
    """Generate invoice data"""
    invoices = []
    completed_appointments = [apt for apt in appointments if apt["status"] == "completed"]
    
    for idx, apt in enumerate(completed_appointments, 1):
        invoice_date = datetime.strptime(apt["appointment_date"], "%Y-%m-%d")
        
        invoices.append({
            "invoice_id": f"INV{idx:05d}",
            "appointment_id": apt["appointment_id"],
            "customer_id": apt["customer_id"],
            "invoice_date": apt["appointment_date"],
            "due_date": (invoice_date + timedelta(days=15)).strftime("%Y-%m-%d"),
            "subtotal": apt["total_amount"],
            "tax_rate": 8.25,
            "tax_amount": round(apt["total_amount"] * 0.0825, 2),
            "discount": random.choice([0, 0, 0, 10, 25, 50]),
            "total_amount": round(apt["total_amount"] * 1.0825, 2),
            "amount_paid": round(apt["total_amount"] * 1.0825, 2),
            "balance": 0,
            "payment_status": "paid",
            "payment_date": apt["appointment_date"],
            "payment_method": apt["payment_method"]
        })
    
    return invoices


def generate_staff():
    """Generate staff/employee data"""
    staff = [
        {
            "staff_id": "STAFF001",
            "first_name": "Beauty",
            "last_name": "Bee",
            "role": "CEO / Lead Makeup Artist",
            "email": "beauty@glamorbybee.com",
            "phone": "(214) 555-0001",
            "hire_date": "2023-01-15",
            "is_active": True,
            "specialties": "Bridal, Special FX, Editorial",
            "hourly_rate": 75.00
        },
        {
            "staff_id": "STAFF002",
            "first_name": "Paul",
            "last_name": "Logistics",
            "role": "Operations Manager",
            "email": "paul@glamorbybee.com",
            "phone": "(214) 555-0002",
            "hire_date": "2023-02-01",
            "is_active": True,
            "specialties": "Scheduling, Inventory, Client Relations",
            "hourly_rate": 45.00
        }
    ]
    return staff


def generate_analytics():
    """Generate analytics/metrics data"""
    analytics = []
    base_date = datetime.now() - timedelta(days=90)
    
    for i in range(90):
        date = base_date + timedelta(days=i)
        # More bookings on weekends
        is_weekend = date.weekday() >= 5
        base_bookings = random.randint(8, 15) if is_weekend else random.randint(3, 8)
        
        analytics.append({
            "date": date.strftime("%Y-%m-%d"),
            "total_bookings": base_bookings,
            "completed_bookings": int(base_bookings * random.uniform(0.85, 0.95)),
            "cancelled_bookings": random.randint(0, 2),
            "no_shows": random.randint(0, 1),
            "total_revenue": round(base_bookings * random.uniform(150, 300), 2),
            "new_clients": random.randint(0, 3),
            "returning_clients": base_bookings - random.randint(0, 3),
            "average_service_value": round(random.uniform(150, 300), 2),
            "deposit_collected": round(base_bookings * random.uniform(45, 90), 2),
            "tips_received": round(base_bookings * random.uniform(10, 30), 2)
        })
    
    return analytics


def generate_inventory():
    """Generate inventory/supplies data"""
    inventory = [
        {"item_id": "INV001", "item_name": "Foundation - Light", "category": "Base", "quantity": 12, "unit": "bottle", "cost_per_unit": 35.00, "reorder_level": 5, "supplier": "MAC Cosmetics"},
        {"item_id": "INV002", "item_name": "Foundation - Medium", "category": "Base", "quantity": 15, "unit": "bottle", "cost_per_unit": 35.00, "reorder_level": 5, "supplier": "MAC Cosmetics"},
        {"item_id": "INV003", "item_name": "Foundation - Dark", "category": "Base", "quantity": 8, "unit": "bottle", "cost_per_unit": 35.00, "reorder_level": 5, "supplier": "MAC Cosmetics"},
        {"item_id": "INV004", "item_name": "Eyeshadow Palette - Neutral", "category": "Eyes", "quantity": 6, "unit": "palette", "cost_per_unit": 52.00, "reorder_level": 3, "supplier": "Urban Decay"},
        {"item_id": "INV005", "item_name": "Eyeshadow Palette - Bold", "category": "Eyes", "quantity": 4, "unit": "palette", "cost_per_unit": 52.00, "reorder_level": 2, "supplier": "Urban Decay"},
        {"item_id": "INV006", "item_name": "Mascara - Waterproof", "category": "Eyes", "quantity": 20, "unit": "tube", "cost_per_unit": 24.00, "reorder_level": 8, "supplier": "Benefit"},
        {"item_id": "INV007", "item_name": "Eyeliner - Black", "category": "Eyes", "quantity": 15, "unit": "pencil", "cost_per_unit": 18.00, "reorder_level": 6, "supplier": "Stila"},
        {"item_id": "INV008", "item_name": "Lipstick - Red", "category": "Lips", "quantity": 10, "unit": "tube", "cost_per_unit": 28.00, "reorder_level": 4, "supplier": "NARS"},
        {"item_id": "INV009", "item_name": "Lipstick - Nude", "category": "Lips", "quantity": 12, "unit": "tube", "cost_per_unit": 28.00, "reorder_level": 5, "supplier": "NARS"},
        {"item_id": "INV010", "item_name": "Lip Gloss", "category": "Lips", "quantity": 18, "unit": "tube", "cost_per_unit": 22.00, "reorder_level": 7, "supplier": "Fenty Beauty"},
        {"item_id": "INV011", "item_name": "Blush - Pink", "category": "Cheeks", "quantity": 8, "unit": "compact", "cost_per_unit": 32.00, "reorder_level": 3, "supplier": "NARS"},
        {"item_id": "INV012", "item_name": "Highlighter", "category": "Cheeks", "quantity": 6, "unit": "compact", "cost_per_unit": 38.00, "reorder_level": 3, "supplier": "Fenty Beauty"},
        {"item_id": "INV013", "item_name": "Bronzer", "category": "Cheeks", "quantity": 7, "unit": "compact", "cost_per_unit": 35.00, "reorder_level": 3, "supplier": "Benefit"},
        {"item_id": "INV014", "item_name": "Setting Spray", "category": "Setting", "quantity": 25, "unit": "bottle", "cost_per_unit": 18.00, "reorder_level": 10, "supplier": "Urban Decay"},
        {"item_id": "INV015", "item_name": "Primer", "category": "Base", "quantity": 14, "unit": "bottle", "cost_per_unit": 32.00, "reorder_level": 6, "supplier": "Smashbox"},
        {"item_id": "INV016", "item_name": "False Lashes - Natural", "category": "Lashes", "quantity": 50, "unit": "pair", "cost_per_unit": 8.00, "reorder_level": 20, "supplier": "Ardell"},
        {"item_id": "INV017", "item_name": "False Lashes - Dramatic", "category": "Lashes", "quantity": 35, "unit": "pair", "cost_per_unit": 12.00, "reorder_level": 15, "supplier": "Huda Beauty"},
        {"item_id": "INV018", "item_name": "Lash Glue", "category": "Lashes", "quantity": 30, "unit": "bottle", "cost_per_unit": 6.00, "reorder_level": 15, "supplier": "DUO"},
        {"item_id": "INV019", "item_name": "Makeup Brushes Set", "category": "Tools", "quantity": 8, "unit": "set", "cost_per_unit": 85.00, "reorder_level": 3, "supplier": "Morphe"},
        {"item_id": "INV020", "item_name": "Beauty Sponges", "category": "Tools", "quantity": 45, "unit": "piece", "cost_per_unit": 6.00, "reorder_level": 20, "supplier": "Beauty Blender"},
        {"item_id": "INV021", "item_name": "Makeup Remover Wipes", "category": "Supplies", "quantity": 60, "unit": "pack", "cost_per_unit": 4.50, "reorder_level": 25, "supplier": "Neutrogena"},
        {"item_id": "INV022", "item_name": "Cotton Pads", "category": "Supplies", "quantity": 100, "unit": "pack", "cost_per_unit": 3.00, "reorder_level": 40, "supplier": "Generic"},
        {"item_id": "INV023", "item_name": "Sanitizing Spray", "category": "Supplies", "quantity": 12, "unit": "bottle", "cost_per_unit": 8.50, "reorder_level": 5, "supplier": "Cinema Secrets"},
    ]
    
    for item in inventory:
        item["total_value"] = round(item["quantity"] * item["cost_per_unit"], 2)
        item["needs_reorder"] = item["quantity"] <= item["reorder_level"]
        item["last_restocked"] = (datetime.now() - timedelta(days=random.randint(5, 60))).strftime("%Y-%m-%d")
    
    return inventory


def generate_schedule():
    """Generate schedule/availability data"""
    schedule = []
    base_date = datetime.now()
    
    # Generate schedule for next 30 days
    for i in range(30):
        date = base_date + timedelta(days=i)
        day_name = date.strftime("%A")
        
        # Skip some days (closed)
        if day_name == "Sunday" or (day_name == "Monday" and random.random() < 0.3):
            schedule.append({
                "date": date.strftime("%Y-%m-%d"),
                "day_of_week": day_name,
                "is_available": False,
                "open_time": None,
                "close_time": None,
                "total_slots": 0,
                "booked_slots": 0,
                "available_slots": 0,
                "staff_assigned": None,
                "notes": "Closed"
            })
        else:
            # Regular business hours
            open_time = "09:00" if day_name == "Saturday" else "10:00"
            close_time = "18:00"
            total_slots = 8
            booked = random.randint(2, 7)
            
            schedule.append({
                "date": date.strftime("%Y-%m-%d"),
                "day_of_week": day_name,
                "is_available": True,
                "open_time": open_time,
                "close_time": close_time,
                "total_slots": total_slots,
                "booked_slots": booked,
                "available_slots": total_slots - booked,
                "staff_assigned": "Beauty Bee",
                "notes": "Special event" if random.random() < 0.1 else ""
            })
    
    return schedule


def generate_settings():
    """Generate business settings/configuration data"""
    settings = [
        {"setting_id": "SET001", "category": "Business", "setting_key": "business_name", "setting_value": "GlamorByBee", "description": "Business name"},
        {"setting_id": "SET002", "category": "Business", "setting_key": "business_email", "setting_value": "contact@glamorbybee.com", "description": "Primary business email"},
        {"setting_id": "SET003", "category": "Business", "setting_key": "business_phone", "setting_value": "(214) 555-0100", "description": "Primary business phone"},
        {"setting_id": "SET004", "category": "Business", "setting_key": "business_address", "setting_value": "123 Beauty Lane, Dallas, TX 75201", "description": "Studio address"},
        {"setting_id": "SET005", "category": "Booking", "setting_key": "default_deposit_percentage", "setting_value": "30", "description": "Default deposit percentage"},
        {"setting_id": "SET006", "category": "Booking", "setting_key": "cancellation_notice_hours", "setting_value": "24", "description": "Cancellation notice required (hours)"},
        {"setting_id": "SET007", "category": "Booking", "setting_key": "max_advance_booking_days", "setting_value": "90", "description": "Maximum days in advance for booking"},
        {"setting_id": "SET008", "category": "Booking", "setting_key": "time_slot_duration", "setting_value": "60", "description": "Default time slot duration (minutes)"},
        {"setting_id": "SET009", "category": "Financial", "setting_key": "tax_rate", "setting_value": "8.25", "description": "Sales tax rate (percentage)"},
        {"setting_id": "SET010", "category": "Financial", "setting_key": "late_fee_percentage", "setting_value": "5", "description": "Late payment fee percentage"},
        {"setting_id": "SET011", "category": "Financial", "setting_key": "invoice_due_days", "setting_value": "15", "description": "Payment due days after service"},
        {"setting_id": "SET012", "category": "Notifications", "setting_key": "send_confirmation_email", "setting_value": "true", "description": "Send booking confirmation emails"},
        {"setting_id": "SET013", "category": "Notifications", "setting_key": "send_reminder_sms", "setting_value": "true", "description": "Send appointment reminder SMS"},
        {"setting_id": "SET014", "category": "Notifications", "setting_key": "reminder_hours_before", "setting_value": "24", "description": "Send reminder X hours before"},
        {"setting_id": "SET015", "category": "Operations", "setting_key": "buffer_time_minutes", "setting_value": "15", "description": "Buffer time between appointments"},
        {"setting_id": "SET016", "category": "Operations", "setting_key": "max_daily_bookings", "setting_value": "8", "description": "Maximum bookings per day"},
    ]
    return settings


def create_excel_file(filename="master.xlsx"):
    """Create Excel file with all data"""
    if not EXCEL_AVAILABLE:
        print("Creating JSON files instead...")
        return create_json_files()
    
    print("Generating dummy data...")
    customers = generate_customers(50)
    addresses = generate_addresses(customers)
    appointments = generate_appointments(customers, 100)
    products = generate_products()
    invoices = generate_invoices(appointments)
    staff = generate_staff()
    analytics = generate_analytics()
    inventory = generate_inventory()
    schedule = generate_schedule()
    settings = generate_settings()
    
    print(f"Creating Excel workbook: {filename}")
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define header style
    header_fill = PatternFill(start_color="D63384", end_color="D63384", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create sheets
    sheets_data = [
        ("Customers", customers),
        ("Addresses", addresses),
        ("Appointments", appointments),
        ("Products", products),
        ("Invoices", invoices),
        ("Staff", staff),
        ("Analytics", analytics),
        ("Inventory", inventory),
        ("Schedule", schedule),
        ("Settings", settings)
    ]
    
    for sheet_name, data in sheets_data:
        print(f"Creating sheet: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        
        if not data:
            continue
        
        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, record in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=record[header])
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(filename)
    print(f"\n✅ Excel file created successfully: {filename}")
    print(f"\nData Summary:")
    print(f"  • Customers: {len(customers)}")
    print(f"  • Addresses: {len(addresses)}")
    print(f"  • Appointments: {len(appointments)}")
    print(f"  • Products/Services: {len(products)}")
    print(f"  • Invoices: {len(invoices)}")
    print(f"  • Staff: {len(staff)}")
    print(f"  • Analytics: {len(analytics)}")
    print(f"  • Inventory: {len(inventory)}")
    print(f"  • Schedule: {len(schedule)}")
    print(f"  • Settings: {len(settings)}")


def create_json_files():
    """Create JSON files as fallback"""
    print("Generating dummy data...")
    customers = generate_customers(50)
    addresses = generate_addresses(customers)
    appointments = generate_appointments(customers, 100)
    products = generate_products()
    invoices = generate_invoices(appointments)
    staff = generate_staff()
    analytics = generate_analytics()
    inventory = generate_inventory()
    schedule = generate_schedule()
    settings = generate_settings()
    
    data = {
        "customers": customers,
        "addresses": addresses,
        "appointments": appointments,
        "products": products,
        "invoices": invoices,
        "staff": staff,
        "analytics": analytics,
        "inventory": inventory,
        "schedule": schedule,
        "settings": settings
    }
    
    # Save individual JSON files
    for key, value in data.items():
        filename = f"../json/{key}.json"
        with open(filename, 'w') as f:
            json.dump({key: value}, f, indent=2)
        print(f"✅ Created: {filename}")
    
    # Save master JSON file
    with open("../json/master_data.json", 'w') as f:
        json.dump(data, f, indent=2)
    print(f"✅ Created: ../json/master_data.json")
    
    print(f"\nData Summary:")
    print(f"  • Customers: {len(customers)}")
    print(f"  • Addresses: {len(addresses)}")
    print(f"  • Appointments: {len(appointments)}")
    print(f"  • Products/Services: {len(products)}")
    print(f"  • Invoices: {len(invoices)}")
    print(f"  • Staff: {len(staff)}")


if __name__ == "__main__":
    create_excel_file("master.xlsx")
